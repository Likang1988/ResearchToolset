import sys
import os
import json
import csv
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from PySide6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
                              QTreeWidget, QTreeWidgetItem, QPushButton, QFileDialog, QMenu)
from PySide6.QtCore import Qt
from PySide6.QtGui import QAction, QIcon

class TreeListApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle('树形列表工具')
        self.resize(800, 600)
        
        # 设置应用图标
        icon_path = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'assets', 'logo', 'treelist.svg')
        icon = QIcon(icon_path)
        self.setWindowIcon(icon)
        
        # 主布局
        main_widget = QWidget()
        self.setCentralWidget(main_widget)
        main_layout = QVBoxLayout(main_widget)
        
        # 按钮栏
        button_layout = QHBoxLayout()
        
        # 左侧按钮组
        left_btns = QHBoxLayout()
        self.btn_create = QPushButton('创建项目')
        self.btn_add_sibling = QPushButton('增加同级')
        self.btn_add_child = QPushButton('增加子级')
        self.btn_delete = QPushButton('删除该级')
        
        # 初始化按钮状态
        self.btn_add_sibling.setEnabled(False)
        self.btn_add_child.setEnabled(False)
        self.btn_delete.setEnabled(False)
        left_btns.addWidget(self.btn_create)
        left_btns.addWidget(self.btn_add_sibling)
        left_btns.addWidget(self.btn_add_child)
        left_btns.addWidget(self.btn_delete)
        self.btn_add_column = QPushButton('增加新列')
        left_btns.addWidget(self.btn_add_column)
        
        # 右侧按钮组
        right_btns = QHBoxLayout()
        self.btn_import = QPushButton('导入')
        self.btn_export = QPushButton('导出')
        right_btns.addWidget(self.btn_import)
        right_btns.addWidget(self.btn_export)
        
        button_layout.addLayout(left_btns)
        button_layout.addStretch()
        button_layout.addLayout(right_btns)
        
        # 树形列表
        self.tree = QTreeWidget()
        self.tree.setHeaderHidden(False)
        self.tree.setColumnCount(1)
        self.tree.setHeaderLabels(['名称'])
        self.tree.setContextMenuPolicy(Qt.CustomContextMenu)
        self.tree.customContextMenuRequested.connect(self.show_context_menu)
        self.tree.itemSelectionChanged.connect(self.update_buttons_state)
        
        main_layout.addLayout(button_layout)
        main_layout.addWidget(self.tree)
        
        # 应用样式
        self.setStyleSheet('''
            QMainWindow {
                background-color: #f5f5f5;
                font-family: "Microsoft YaHei";
            }
            QTreeWidget {
                background-color: transparent;
                border: 1px solid rgba(0, 0, 0, 0.1);
                border-radius: 8px;
                selection-background-color: rgba(0, 120, 212, 0.1);
                selection-color: black;
            }
            QTreeWidget::item {
                height: 36px;
                padding: 2px;
            }
            QTreeWidget::item:hover {
                background-color: rgba(0, 0, 0, 0.05);
            }
            QTreeWidget::item:selected {
                background-color: rgba(0, 120, 212, 0.1);
                color: black;
            }
            QTreeWidget::item:selected:active {
                background-color: rgba(0, 120, 212, 0.15);
                color: black;
            }
            QTreeWidget QHeaderView::section {
                background-color: #f3f3f3;
                color: #333333;
                font-weight: 500;
                padding: 8px;
                border: none;
                border-bottom: 1px solid rgba(0, 0, 0, 0.1);
            }
            QHeaderView::section:hover {
                background-color: #e5e5e5;
            }
            
            QPushButton {
                padding: 8px 16px;
                border-radius: 4px;
                background-color: #4CAF50;
                color: white;
                border: none;
                font-size: 14px;
            }
            QPushButton:hover {
                background-color: #45a049;
            }
            QPushButton:pressed {
                background-color: #388e3c;
            }
            QPushButton:disabled {
                background-color: #cccccc;
                color: #888888;
            }
        ''')
        
        # 连接信号
        self.btn_create.clicked.connect(self.create_root_item)
        self.btn_add_sibling.clicked.connect(self.add_sibling_item)
        self.btn_add_child.clicked.connect(self.add_child_item)
        self.btn_delete.clicked.connect(self.delete_item)
        self.btn_add_column.clicked.connect(self.add_column_item)
        self.btn_import.clicked.connect(self.import_data)
        self.btn_export.clicked.connect(self.export_data)
        
    def import_data(self):
        path, _ = QFileDialog.getOpenFileName(self, '导入文件', '', 'JSON文件 (*.json)')
        if path:
            try:
                with open(path, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                # 清空当前树
                self.tree.clear()
                # 设置列数
                if 'columns' in data and len(data['columns']) > 0:
                    self.tree.setColumnCount(len(data['columns']))
                    self.tree.setHeaderLabels(data['columns'])
                # 递归构建树
                self._build_tree_from_data(data, self.tree.invisibleRootItem())
            except Exception as e:
                from PySide6.QtWidgets import QMessageBox
                QMessageBox.critical(self, '导入错误', f'导入文件时发生错误：{str(e)}')
    
    def _build_tree_from_data(self, data, parent_item):
        """从JSON数据递归构建树"""
        # 处理子节点
        for child_data in data.get('children', []):
            child = QTreeWidgetItem(parent_item)
            # 设置各列的值
            for col, text in enumerate(child_data.get('columns', [])):
                child.setText(col, text)
            # 设置可编辑
            child.setFlags(child.flags() | Qt.ItemIsEditable)
            # 递归处理子节点
            self._build_tree_from_data(child_data, child)

    def export_data(self):
        path, filter = QFileDialog.getSaveFileName(self, '导出文件', '',
                                                 'Excel文件 (*.xlsx);;CSV文件 (*.csv);;JSON文件 (*.json)',
                                                 options=QFileDialog.Options(),
                                                 selectedFilter='Excel文件 (*.xlsx)')
        if path:
            if filter == 'JSON文件 (*.json)':
                # 使用原有的保存JSON功能
                data = self._get_tree_data(self.tree.invisibleRootItem())
                with open(path, 'w', encoding='utf-8') as f:
                    json.dump(data, f, ensure_ascii=False, indent=2)
            else:
                items = self._flatten_tree(self.tree.invisibleRootItem())
                if filter == 'CSV文件 (*.csv)':
                    self._export_csv(items, path)
                else:
                    self._export_excel(items, path)

    def _get_tree_data(self, item):
        data = {
            'columns': [item.text(col) for col in range(self.tree.columnCount())],
            'children': []
        }
        for i in range(item.childCount()):
            child = item.child(i)
            data['children'].append(self._get_tree_data(child))
        return data

    def _flatten_tree(self, item, level=1, parent_chain=[]):
        items = []
        
        # 检查是否为invisibleRootItem
        if item == self.tree.invisibleRootItem():
            # 如果是invisibleRootItem，只处理其子项
            for i in range(item.childCount()):
                child = item.child(i)
                items.extend(self._flatten_tree(child, level, parent_chain))
            return items
        
        current_chain = parent_chain + [item.text(0)]
        
        # 生成层级列数据
        item_data = {}
        for i, name in enumerate(current_chain):
            item_data[f'level_{i}'] = name
        
        # 保留自定义列数据（从第二列开始）
        for col in range(1, self.tree.columnCount()):
            item_data[f'col_{col}'] = item.text(col) if item.text(col) else ''
            
        # 如果是叶子节点或者最末级节点，添加到结果中
        if item.childCount() == 0:
            items.append(item_data)
        else:
            # 如果有子节点，递归处理子节点
            for i in range(item.childCount()):
                child = item.child(i)
                items.extend(self._flatten_tree(child, level+1, current_chain))
            
            # 如果当前节点有自定义列数据，也添加到结果中
            has_custom_data = any(item.text(col) for col in range(1, self.tree.columnCount()))
            if has_custom_data:
                items.append(item_data)
                
        return items

    def _get_max_depth(self):
        def get_depth(item):
            if not item.childCount():
                return 1
            return 1 + max(get_depth(item.child(i)) for i in range(item.childCount()))
        
        root = self.tree.invisibleRootItem()
        return max((get_depth(root.child(i)) for i in range(root.childCount())), default=0)

    def _export_csv(self, items, path):
        with open(path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            
            # 动态生成表头
            max_depth = self._get_max_depth()
            level_headers = [f'层级{i+1}' for i in range(max_depth)]
            custom_headers = [self.tree.headerItem().text(col) for col in range(1, self.tree.columnCount())]
            writer.writerow(level_headers + custom_headers)
            
            for item in items:
                row = [item.get(f'level_{i}', '') for i in range(max_depth)]
                row.extend(item.get(f'col_{col}', '') for col in range(1, self.tree.columnCount()))
                writer.writerow(row)

    def _export_excel(self, items, path):
        wb = Workbook()
        ws = wb.active
        
        # 动态生成层级列表头
        max_depth = self._get_max_depth()
        level_headers = [f'层级{i+1}' for i in range(max_depth)]
        custom_headers = [self.tree.headerItem().text(col) for col in range(1, self.tree.columnCount())]
        
        # 组合表头
        headers = level_headers + custom_headers
        ws.append(headers)
        
        # 设置表头样式
        header_fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), 
            right=Side(style='thin'), 
            top=Side(style='thin'), 
            bottom=Side(style='thin')
        )
        
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border
            # 设置列宽
            ws.column_dimensions[get_column_letter(col_idx)].width = 15
        
        # 构建层级关系数据结构
        hierarchy_data = self._build_hierarchy_data(items, max_depth)
        
        # 填充数据并合并单元格
        row_idx = 2  # 从第2行开始（表头是第1行）
        self._fill_excel_data(ws, hierarchy_data, row_idx, 1, max_depth, thin_border)
        
        wb.save(path)
    
    def _build_hierarchy_data(self, items, max_depth):
        """构建层级关系数据结构"""
        # 初始化层级结构
        hierarchy = {}
        
        for item in items:
            current = hierarchy
            # 遍历每个层级
            for level in range(max_depth):
                level_name = item.get(f'level_{level}', '')
                if not level_name:
                    break
                    
                if level_name not in current:
                    current[level_name] = {'children': {}, 'data': {}}
                    
                # 存储自定义列数据
                if level == max_depth - 1 or not item.get(f'level_{level+1}', ''):
                    for col in range(1, self.tree.columnCount()):
                        current[level_name]['data'][f'col_{col}'] = item.get(f'col_{col}', '')
                        
                current = current[level_name]['children']
                
        return hierarchy
    
    def _fill_excel_data(self, ws, hierarchy, start_row, level, max_depth, border_style):
        """递归填充Excel数据并合并单元格"""
        current_row = start_row
        
        for name, data in hierarchy.items():
            # 计算此层级项目占用的行数
            children = data['children']
            if not children:  # 叶子节点
                row_span = 1
            else:
                # 计算所有子项占用的总行数
                row_span = 0
                for child_name, child_data in children.items():
                    child_rows = self._count_rows(child_data)
                    row_span += child_rows
            
            # 在当前层级列写入数据
            cell = ws.cell(row=current_row, column=level, value=name)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border_style
            
            # 设置背景色
            cell.fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
            
            # 如果占多行，合并单元格
            if row_span > 1:
                ws.merge_cells(start_row=current_row, start_column=level, 
                              end_row=current_row + row_span - 1, end_column=level)
            
            # 如果是叶子节点，填充自定义列数据
            if not children:
                for col in range(1, self.tree.columnCount()):
                    col_value = data['data'].get(f'col_{col}', '')
                    cell = ws.cell(row=current_row, column=max_depth + col - 1, value=col_value)
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                    cell.border = border_style
            
            # 递归处理子项
            if children:
                child_row = current_row
                for child_name, child_data in children.items():
                    child_row = self._fill_excel_data(ws, {child_name: child_data}, child_row, level + 1, max_depth, border_style)
            
            current_row += row_span
        
        return current_row
    
    def _count_rows(self, data):
        """计算节点及其子节点占用的总行数"""
        children = data['children']
        if not children:  # 叶子节点占一行
            return 1
            
        # 非叶子节点，计算所有子节点的行数总和
        total_rows = 0
        for child_name, child_data in children.items():
            total_rows += self._count_rows(child_data)
            
        return total_rows
    
    def create_root_item(self):
        item = QTreeWidgetItem(self.tree)
        item.setText(0, '请输入项目名称')
        item.setFlags(item.flags() | Qt.ItemIsEditable)
        
    def add_sibling_item(self):
        current = self.tree.currentItem()
        if current:
            parent = current.parent()
            item = QTreeWidgetItem()
            item.setText(0, '请输入此级名称')
            item.setFlags(item.flags() | Qt.ItemIsEditable)
            (parent or self.tree).insertChild((parent or self.tree).indexOfChild(current) + 1, item)
        
    def add_child_item(self):
        current = self.tree.currentItem()
        if current:
            item = QTreeWidgetItem(current)
            item.setText(0, '请输入此级名称')
            item.setFlags(item.flags() | Qt.ItemIsEditable)
            current.setExpanded(True)
        
    def delete_item(self):
        current = self.tree.currentItem()
        if current:
            # 允许删除根节点但保留至少一个根项目
            root = self.tree.invisibleRootItem()
            if current.parent() is None and root.childCount() > 1:
                root.removeChild(current)
            elif current.parent():
                current.parent().removeChild(current)

    def add_column_item(self):
        from PySide6.QtWidgets import QInputDialog
        col_name, ok = QInputDialog.getText(self, '新建列', '请输入列标题:')
        if ok and col_name:
            # 动态增加列
            self.tree.setColumnCount(self.tree.columnCount() + 1)
            self.tree.setHeaderLabels([self.tree.headerItem().text(i) for i in range(self.tree.columnCount()-1)] + [col_name])
            
            # 为所有现有项初始化新列
            root = self.tree.invisibleRootItem()
            self._init_new_column(root)
    
    def _init_new_column(self, item):
        for i in range(item.childCount()):
            child = item.child(i)
            child.setText(self.tree.columnCount()-1, '')
            child.setFlags(child.flags() | Qt.ItemIsEditable)
            self._init_new_column(child)

    def contextMenuEvent(self, event):
        menu = QMenu(self)
        menu.addAction('刷新样式', self._update_style)
        menu.exec_(event.globalPos())
    
    def _update_style(self):
        self.setStyleSheet(self.styleSheet())  # 强制刷新样式
        
    def update_buttons_state(self):
        current = self.tree.currentItem()
        has_selection = current is not None
        
        self.btn_add_sibling.setEnabled(has_selection and current.parent() is not None)
        self.btn_add_child.setEnabled(has_selection)
        self.btn_delete.setEnabled(has_selection)
        
        # 根节点特殊处理
        if current and current.parent() is None:
            self.btn_add_sibling.setEnabled(False)
            self.btn_delete.setEnabled(self.tree.invisibleRootItem().childCount() > 1)
        
    def show_context_menu(self, pos):
        menu = QMenu()
        rename_action = QAction('重命名', self)
        rename_action.triggered.connect(lambda: self.tree.editItem(self.tree.currentItem()))
        menu.addAction(rename_action)
        menu.exec_(self.tree.viewport().mapToGlobal(pos))

if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = TreeListApp()
    window.show()
    sys.exit(app.exec())